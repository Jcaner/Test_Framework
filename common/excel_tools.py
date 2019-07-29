import xlrd
from openpyxl import load_workbook
from config.config import Config


class ExcelTools(object):

    def __init__(self, rw, file, index=0):
        if rw == 'r':
            self.wbr = xlrd.open_workbook(filename=file)
            self.sheet = self.wbr.sheet_by_index(index)
            self.rows = self.sheet.nrows
            self.cols = self.sheet.ncols
            self.keys = self.sheet.row_values(1)
        elif rw == 'w':
            self.file = file
            self.wbw = load_workbook(filename=file)
            self.ws = self.wbw.active
        else:
            print("参数错误")

    def get_sheet(self):
        return self.sheet

    def get_rows(self):
        return self.rows

    def get_cols(self):
        return self.cols

    def get_cell(self, row, col):
        return self.sheet.cell(row, col)

    def get_cell_type(self, row, col):
        return self.sheet.cell_type(row, col)

    def get_cell_value(self, row, col):
        return self.sheet.cell_value(row, col)

    def read_excel(self):
        for row in range(2, self.sheet.nrows):
            yield self.sheet.row_values(row)

    def dict_values(self):
        if self.rows <= 2:
            print("excel表格无数据")
        else:
            r = []
            j = 2
            for i in range(self.rows - 2):
                s = dict()
                s['CaseId'] = i + 1
                values = self.sheet.row_values(j)
                for x in range(self.cols):
                    s[self.keys[x]] = values[x]
                r.append(s)
                j += 1
            return r

    def write_excel(self, row, col, value):
        self.ws.cell(row, col).value = value
        self.wbw.save(self.file)


if __name__ == "__main__":
    excel_file = Config().base_path + '/testdata/xlsx/test_register_data.xlsx'
    et = ExcelTools('r', excel_file)
    data = et.read_excel()
    print(et.get_rows())
    print(et.get_cols())
    print(et.get_cell_value(2, 6))
    print(et.get_cell_value(3, 6))
    if et.get_cell_value(4, 6) == '':
        print('true')
    print(next(data))
    print(next(data))
    print(next(data))
    print(et.dict_values())
